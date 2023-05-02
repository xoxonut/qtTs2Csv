import openpyxl
from lxml import etree

# 讀取XML檔案並獲取根元素
input_name = 'in'
output_name ='out'
tree = etree.parse(input_name)
root = tree.getroot()

# 創建一個新的XLSX工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 添加標題行
headers = ["content_name", "location_file", "location_line", "source", "translation"]
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num).value = header

# 遍歷<context>標籤並將資料寫入XLSX
row_num = 2
for context in root.findall("context"):
    content_name = context.find("name")
    content_name = content_name.text if content_name is not None else ""

    # 遍歷<message>標籤
    for message in context.findall("message"):
        source = message.find("source")
        source = source.text if source is not None else ""

        translation = message.find("translation")
        translation = translation.text if translation is not None else ""

        locations = message.findall("location")
        if locations:
            for location in locations:
                location_file = location.get("filename") if location is not None else ""
                location_line = location.get("line") if location is not None else ""

                # 寫入XLSX行
                for col_num, cell_value in enumerate([content_name, location_file, location_line, source, translation], 1):
                    worksheet.cell(row=row_num, column=col_num).value = cell_value
                row_num += 1
        else:
            # 如果沒有<location>標籤，則使用空字符串
            location_file = ""
            location_line = ""

            # 寫入XLSX行
            for col_num, cell_value in enumerate([content_name, location_file, location_line, source, translation], 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value
            row_num += 1

# 保存XLSX檔案
workbook.save(output_name)

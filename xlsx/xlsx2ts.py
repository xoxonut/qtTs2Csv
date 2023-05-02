import openpyxl
from lxml import etree

# 读取XML文件并获取根元素
input_name = 'in'
output_name ='out'
parser = etree.XMLParser(strip_cdata=False)
tree = etree.parse(input_name, parser)
root = tree.getroot()

# 读取XLSX文件
workbook = openpyxl.load_workbook("output.xlsx")
worksheet = workbook.active

# 遍历XLSX中的每一行
for row_num in range(2, worksheet.max_row + 1):
    content_name = worksheet.cell(row=row_num, column=1).value
    location_file = worksheet.cell(row=row_num, column=2).value
    location_line = worksheet.cell(row=row_num, column=3).value
    source_text = worksheet.cell(row=row_num, column=4).value
    translation_text = worksheet.cell(row=row_num, column=5).value

    # 找到匹配的<context>标签
    context_match = None
    for context in root.findall("context"):
        name = context.find("name")
        if name is not None and name.text == content_name:
            context_match = context
            break

    if context_match is None:
        continue

    # 找到匹配的<message>标签
    message_match = None
    for message in context_match.findall("message"):
        source = message.find("source")
        if source is not None and source.text == source_text:
            for location in message.findall("location"):
                if (location.get("filename") == location_file and
                        location.get("line") == location_line):
                    message_match = message
                    break

        if message_match is not None:
            break

    if message_match is None:
        continue

    # 更新<translation>标签的内容
    translation = message_match.find("translation")
    if translation is not None:
        translation.text = translation_text

# 将更新后的XML树写回文件，添加xml声明和doctype
with open(output_name, "wb") as f:
    f.write(b'<?xml version="1.0" encoding="utf-8"?>\n')
    f.write(b'<!DOCTYPE TS>\n')
    f.write(etree.tostring(root, encoding="utf-8"))
